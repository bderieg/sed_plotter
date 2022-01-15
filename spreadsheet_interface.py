import openpyxl as opxl
import galaxy


# FOLLOWING: Constants
spreadsheet_loc = "C:\\Users\\bderi\\Box\\School\\Research Boizelle\\ALMA Archive Galaxy Observations with SED Flux Densities.xlsx"

# Gets the constant LINES_PER_DATA_SET
def lines_per_data_set():
    wb = opxl.load_workbook(spreadsheet_loc, data_only=True, read_only=True)
    sheet = wb.get_sheet_by_name("SEDs")
    found = False

    cur_row = 1
    while found is False:
        cur_row += 1
        if sheet.cell(row=cur_row+1, column=1).value is not None:
            found = True

    return cur_row - 2


LINES_PER_DATA_SET = lines_per_data_set()

# FOLLOWING: Write functions

def new_fit(galaxy_name, fit_type, parameters, fit_range, linestyle):
    wb = opxl.load_workbook(spreadsheet_loc)
    sheet = wb.get_sheet_by_name("Fit Parameters")

    found = False
    new_galaxy = False

    # Find galaxy name
    cur_row = 2
    while found == False:
        if sheet.cell(row=cur_row, column=1).value == galaxy_name:
            found = True
        elif sheet.cell(row=cur_row, column=2).value == None:
            found = True
            new_galaxy = True
        cur_row += 1
    cur_row -= 1

    # Insert fit, creating new section for galaxy if it doesn't already exist.
    if new_galaxy == True:
        sheet.insert_rows(cur_row)
        sheet.cell(row=cur_row, column=1).value = galaxy_name
    if sheet.cell(row=cur_row, column=2).value != None:
        cur_row += 1
        sheet.insert_rows(cur_row)
    sheet.cell(row=cur_row, column=2).value = fit_type
    for i in range(3, len(parameters)+3):
        sheet.cell(row=cur_row, column=i).value = parameters[i-3]

    # Insert fit ranges
    for i in range(6, 8):
        sheet.cell(row=cur_row, column=i).value = fit_range[i-6]

    # Insert line_style
    sheet.cell(row=cur_row, column=8).value = linestyle

    wb.save(spreadsheet_loc)

def clear_fits(galaxy_name, fit_type):
    wb = opxl.load_workbook(spreadsheet_loc)
    sheet = wb.get_sheet_by_name("Fit Parameters")

    found = False

    # Find galaxy name
    cur_row = 2
    going_for_a_while = 0
    while found == False:
        if sheet.cell(row=cur_row, column=1).value == galaxy_name:
            found = True
        elif sheet.cell(row=cur_row, column=1).value != None:
            going_for_a_while = 0
        cur_row += 1
        going_for_a_while += 1
        if going_for_a_while > 20:
            return None
    cur_row -= 1

    # Iterate over rows and remove them if they contain the fit type
    name_row = cur_row
    replace = True
    end = False
    while end is False:
        if sheet.cell(row=cur_row, column=2).value == fit_type:
            sheet.delete_rows(cur_row)
        else:
            cur_row += 1
            going_for_a_while += 1
        if (sheet.cell(row=cur_row, column=1).value != None) or (sheet.cell(row=cur_row, column=2).value == None):
            if name_row == cur_row:
                replace = False
            end = True

    # Replace name if it was deleted
    if replace is True:
        sheet.cell(row=name_row, column=1).value = galaxy_name

    wb.save(spreadsheet_loc)

# FOLLOWING: Read functions

def get_set_n(n):  # Assumes LINES_PER_DATA_SET (constant) lines of data per galaxy
    # Puts data from the nth galaxy data set (on sheet named "SEDs") into a Galaxy class (returns Galaxy class)

    wb = opxl.load_workbook(spreadsheet_loc, data_only=True, read_only=True)
    sheet = wb.get_sheet_by_name("SEDs")

    init_row = 2 + (n*(LINES_PER_DATA_SET+1))
    reached_end = False
    col = 3

    # Iterate over the lists of values
    galaxy_name = sheet["A" + str(init_row)].value
    num_lists = 5
    lists = []
    for i in range(num_lists):
        lists.append([])
        while reached_end is False:
            if sheet.cell(row=init_row+i, column=col).value is not None:
                cur_val = sheet.cell(row=init_row+i, column=col).value
                if type(cur_val) is str:
                    if ("None" in cur_val) or ("none" in cur_val):
                        lists[i].append(0.0)
                    else:
                        lists[i].append(cur_val)
                else:
                    lists[i].append(cur_val)
            else:
                reached_end = True
            col += 1
        col = 3
        reached_end = False

    # Get redshift and distance
    z = sheet["C" + str(init_row+5)].value
    distance = sheet["C" + str(init_row+6)].value

    # Create galaxy and return it
    g = galaxy.Galaxy(galaxy_name, lists[0], lists[1], lists[2],
                      lists[3], lists[4], z, distance)
    return g

def read_fits(galaxy_name):
    fit_list = []

    wb = opxl.load_workbook(spreadsheet_loc, data_only=True, read_only=True)
    sheet = wb.get_sheet_by_name("Fit Parameters")

    found = False

    # Find galaxy name, returning an empty list if the loop goes more than 20 iterations without finding the name.
    going_for_a_while = 0
    cur_row = 2
    while found == False:
        if sheet.cell(row=cur_row, column=1).value == galaxy_name:
            found = True
        cur_row += 1
        if sheet.cell(row=cur_row, column=1).value == None:
            going_for_a_while += 1
        else:
            going_for_a_while = 0
        if going_for_a_while > 20:
            return fit_list
    cur_row -= 1

    # Populate fit_list
    end = False
    while end is False:
        cur_list = []
        for i in range(2, 9):
            cur_list.append(sheet.cell(row=cur_row, column=i).value)
        fit_list.append(cur_list)
        if (sheet.cell(row=cur_row + 1, column=1).value != None) or (sheet.cell(row=cur_row + 1, column=2).value == None):
            end = True
        cur_row += 1

    if fit_list[0][0] == None:
        return []
    else:
        return fit_list

# Returns a list of targets based on the SEDs spreadsheet
def target_list():
    wb = opxl.load_workbook(spreadsheet_loc, data_only=True, read_only=True)
    sheet = wb.get_sheet_by_name("SEDs")

    name_list = []

    loop = True
    cur_row = 2
    while loop is True:
        name_list.append(sheet.cell(row=cur_row, column=1).value)
        cur_row += LINES_PER_DATA_SET + 1
        if sheet.cell(row=cur_row, column=1).value is None:
            loop = False

    return name_list

def get_point_styles():
    wb = opxl.load_workbook(spreadsheet_loc, data_only=True, read_only=True)
    sheet = wb.get_sheet_by_name("Point Styles")

    # Define first row and column
    row_start = 1
    column_start = 2

    # Populate the list to return, going from top to bottom
    column = column_start
    end_row = False
    comp_list = []
    for row in range(row_start, row_start+3):
        comp_list.append([])
        while end_row is False:
            comp_list[row-row_start].append(sheet.cell(row=row, column=column).value)
            if sheet.cell(row=row, column=column+1).value == None:
                end_row = True
            column += 1
        end_row = False
        column = column_start

    return comp_list
